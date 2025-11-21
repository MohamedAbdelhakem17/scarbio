# -*- coding: utf-8 -*-
"""
============================================================================
KEYWORD ANALYZER - ENCODING SAFE VERSION
============================================================================
"""

import pandas as pd
import json
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from bs4 import BeautifulSoup
import warnings
import sys
import os
import io

warnings.filterwarnings('ignore')

class SilentKeywordAnalyzer:
    def __init__(self, input_file, filter_option='recommended', output_dir='results'):
        self.input_file = input_file
        self.filter_option = filter_option
        self.output_dir = output_dir
        self.data_source = None
        self.onpage_results = []
        self.consolidated_results = []
        
        self.colors = {
            'header': '9b5277',
            'yes': 'bdd6ac',
            'no': 'c58c8a',
            'primary_keyword': 'ffd966',
            'secondary_keywords': 'c9daf8'
        }
    
    def detect_data_source(self, df):
        columns = [col.lower() for col in df.columns]
        
        if 'keyword' in columns and 'url' in columns and 'search volume' in columns:
            return 'semrush'
        elif 'query' in columns and 'page' in columns:
            return 'gsc'
        else:
            return 'unknown'
    
    def normalize_dataframe(self, df):
        if self.data_source == 'semrush':
            column_mapping = {}
            for col in df.columns:
                col_lower = col.lower()
                if col_lower == 'keyword':
                    column_mapping[col] = 'Query'
                elif col_lower == 'url':
                    column_mapping[col] = 'Page'
                elif col_lower == 'search volume':
                    column_mapping[col] = 'Search Volume'
                elif col_lower == 'position' and 'previous' not in col_lower:
                    column_mapping[col] = 'Position'
            
            df = df.rename(columns=column_mapping)
            essential_cols = ['Query', 'Page', 'Position', 'Search Volume']
            available_cols = [col for col in essential_cols if col in df.columns]
            cols_to_keep = available_cols.copy()
            df = df[cols_to_keep]
            
            if 'Search Volume' in df.columns:
                df['Search Volume'] = pd.to_numeric(df['Search Volume'], errors='coerce').fillna(0)
                df['Impressions'] = df['Search Volume']
                df['Clicks'] = 0
                df['Traffic Opportunity'] = df['Search Volume']
                df['CTR'] = 'N/A'
            
        elif self.data_source == 'gsc':
            if 'Clicks' in df.columns and 'Impressions' in df.columns:
                df['Clicks'] = pd.to_numeric(df['Clicks'], errors='coerce').fillna(0)
                df['Impressions'] = pd.to_numeric(df['Impressions'], errors='coerce').fillna(0)
                df['Traffic Opportunity'] = df['Impressions'] - df['Clicks']
                df['Traffic Opportunity'] = df['Traffic Opportunity'].clip(lower=0)
        
        return df
    
    def load_data(self):
        try:
            if self.input_file.endswith('.csv'):
                df = pd.read_csv(self.input_file, encoding='utf-8')
            elif self.input_file.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(self.input_file)
            else:
                return None
            
            self.data_source = self.detect_data_source(df)
            
            if self.data_source == 'unknown':
                return None
            
            df = self.normalize_dataframe(df)
            
            required_columns = ['Query', 'Page']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return None
            
            has_position = 'Position' in df.columns
            
            if has_position:
                df['Position'] = pd.to_numeric(df['Position'], errors='coerce')
                
                if self.filter_option == 'recommended':
                    df = df[(df['Position'] >= 5) & (df['Position'] <= 20)]
            
            df = df.drop_duplicates(subset=['Query', 'Page'], keep='first')
            
            return df
            
        except Exception as e:
            return None
    
    def get_page_content(self, url, timeout=30000):
        try:
            from playwright.sync_api import sync_playwright
            
            with sync_playwright() as p:
                browser = p.chromium.launch(
                    headless=True,
                    args=[
                        '--disable-blink-features=AutomationControlled',
                        '--no-sandbox',
                        '--disable-dev-shm-usage'
                    ]
                )
                
                context = browser.new_context(
                    viewport={'width': 1920, 'height': 1080},
                    user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
                )
                
                page = context.new_page()
                
                page.add_init_script("""
                    Object.defineProperty(navigator, 'webdriver', {
                        get: () => undefined
                    });
                """)
                
                page.goto(url, wait_until='networkidle', timeout=timeout)
                time.sleep(2)
                page.wait_for_selector('body', timeout=5000)
                
                html_content = page.content()
                browser.close()
                
                return BeautifulSoup(html_content, 'html.parser')
            
        except Exception:
            try:
                from playwright.sync_api import sync_playwright
                
                with sync_playwright() as p:
                    browser = p.chromium.launch(headless=True)
                    context = browser.new_context(
                        user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                    )
                    page = context.new_page()
                    page.goto(url, wait_until='domcontentloaded', timeout=60000)
                    time.sleep(3)
                    html_content = page.content()
                    browser.close()
                    
                    return BeautifulSoup(html_content, 'html.parser')
                    
            except Exception:
                return None
    
    def check_keyword_presence(self, soup, keyword):
        if not soup:
            return {k: 'NO' for k in ['title', 'description', 'h1', 'h2', 'content', 'alt_text', 'schema']}
        
        keyword_lower = keyword.lower()
        results = {}
        
        try:
            title_tag = soup.find('title')
            results['title'] = 'YES' if (title_tag and keyword_lower in title_tag.get_text().lower()) else 'NO'
            
            meta_desc = soup.find('meta', attrs={'name': 'description'})
            results['description'] = 'YES' if (meta_desc and keyword_lower in meta_desc.get('content', '').lower()) else 'NO'
            
            h1_tags = soup.find_all('h1')
            h1_text = ' '.join([tag.get_text().lower() for tag in h1_tags])
            results['h1'] = 'YES' if keyword_lower in h1_text else 'NO'
            
            h2_tags = soup.find_all('h2')
            h2_text = ' '.join([tag.get_text().lower() for tag in h2_tags])
            results['h2'] = 'YES' if keyword_lower in h2_text else 'NO'
            
            p_tags = soup.find_all('p')
            content_text = ' '.join([tag.get_text().lower() for tag in p_tags])
            results['content'] = 'YES' if keyword_lower in content_text else 'NO'
            
            img_tags = soup.find_all('img', alt=True)
            alt_texts = ' '.join([img.get('alt', '').lower() for img in img_tags])
            results['alt_text'] = 'YES' if keyword_lower in alt_texts else 'NO'
            
            schema_scripts = soup.find_all('script', type='application/ld+json')
            schema_text = ''
            for script in schema_scripts:
                try:
                    schema_data = json.loads(script.string)
                    schema_text += json.dumps(schema_data).lower() + ' '
                except Exception:
                    continue
            results['schema'] = 'YES' if keyword_lower in schema_text else 'NO'
            
        except Exception:
            results = {k: 'NO' for k in ['title', 'description', 'h1', 'h2', 'content', 'alt_text', 'schema']}
        
        return results
    
    def run_onpage_analysis(self, df):
        unique_urls = df['Page'].unique()
        url_cache = {}
        total = len(df)
        
        for idx, row in df.iterrows():
            keyword = row['Query']
            page_url = row['Page']
            
            if self.data_source == 'semrush':
                search_volume = row.get('Search Volume', 'N/A')
                traffic_opportunity = row.get('Traffic Opportunity', 'N/A')
                position = row.get('Position', 'N/A')
                clicks = 'N/A'
                impressions = 'N/A'
                ctr = 'N/A'
            else:
                clicks = row.get('Clicks', 'N/A')
                impressions = row.get('Impressions', 'N/A')
                ctr = row.get('CTR', 'N/A')
                position = row.get('Position', 'N/A')
                traffic_opportunity = row.get('Traffic Opportunity', 'N/A')
                search_volume = 'N/A'
            
            if page_url in url_cache:
                soup = url_cache[page_url]
            else:
                soup = self.get_page_content(page_url)
                url_cache[page_url] = soup
                delay = 2.0 if soup else 5.0
                time.sleep(delay)
            
            keyword_presence = self.check_keyword_presence(soup, keyword)
            
            result = {
                'Keyword': keyword,
                'URL': page_url,
                'Title': keyword_presence['title'],
                'Description': keyword_presence['description'],
                'H1': keyword_presence['h1'],
                'H2': keyword_presence['h2'],
                'Content': keyword_presence['content'],
                'Alt Text': keyword_presence['alt_text'],
                'Schema': keyword_presence['schema']
            }
            
            if self.data_source == 'semrush':
                result['Traffic Opportunity'] = traffic_opportunity
                result['Position'] = position
            else:
                result['Clicks'] = clicks
                result['Impressions'] = impressions
                result['CTR'] = ctr
                result['Position'] = position
                result['Traffic Opportunity'] = traffic_opportunity
            
            self.onpage_results.append(result)
    
    def create_url_keyword_mapping(self, results):
        df = pd.DataFrame(results)
        
        if self.data_source == 'semrush':
            metric_col = 'Traffic Opportunity'
            df[metric_col] = pd.to_numeric(df.get(metric_col, 0), errors='coerce').fillna(0)
        else:
            metric_col = 'Clicks'
            df[metric_col] = pd.to_numeric(df.get(metric_col, 0), errors='coerce').fillna(0)
        
        grouped = df.groupby('URL')
        mapping_results = []
        
        for url, group in grouped:
            group_sorted = group.sort_values(metric_col, ascending=False)
            main_row = group_sorted.iloc[0]
            main_keyword = main_row['Keyword']
            secondary = group_sorted.iloc[1:7]['Keyword'].tolist()
            secondary_str = '; '.join(secondary) if secondary else 'None'
            
            mapping_results.append({
                'URL': url,
                'Main Keyword': main_keyword,
                'Secondary Keywords (Top 6)': secondary_str
            })
        
        mapping_df = pd.DataFrame(mapping_results)
        return mapping_df
    
    def create_excel_report(self, filename=None):
        if not filename:
            source_prefix = 'semrush' if self.data_source == 'semrush' else 'gsc'
            filter_suffix = 'pos5-20' if self.filter_option == 'recommended' else 'all'
            filename = f'keyword_analysis_{source_prefix}_{filter_suffix}.xlsx'
        
        os.makedirs(self.output_dir, exist_ok=True)
        filepath = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        wb.remove(wb.active)
        
        if self.onpage_results:
            ws1 = wb.create_sheet("On-Page Targeting")
            df1 = pd.DataFrame(self.onpage_results)
            
            audit_cols = ['Keyword', 'URL', 'Title', 'Description', 'H1', 'H2', 'Content', 'Alt Text', 'Schema']
            
            if self.data_source == 'semrush':
                metric_cols = ['Traffic Opportunity', 'Position']
            else:
                metric_cols = ['Clicks', 'Impressions', 'CTR', 'Position', 'Traffic Opportunity']
            
            final_cols = audit_cols + metric_cols
            df1 = df1[[col for col in final_cols if col in df1.columns]]
            
            for r in dataframe_to_rows(df1, index=False, header=True):
                ws1.append(r)
            
            self.format_onpage_sheet(ws1)
        
        if self.onpage_results:
            mapping_df = self.create_url_keyword_mapping(self.onpage_results)
            ws2 = wb.create_sheet("Keyword Mapping")
            for r in dataframe_to_rows(mapping_df, index=False, header=True):
                ws2.append(r)
            
            self.format_mapping_sheet(ws2)
        
        wb.save(filepath)
        return filename
    
    def format_mapping_sheet(self, ws):
        header_fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        main_fill = PatternFill(start_color='ffd966', end_color='ffd966', fill_type='solid')
        secondary_fill = PatternFill(start_color='c9daf8', end_color='c9daf8', fill_type='solid')
        
        header_font = Font(color='FFFFFF', bold=True)
        regular_font = Font(color='000000')
        bold_font = Font(color='000000', bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = regular_font
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                col_header = ws.cell(row=1, column=cell.column).value
                
                if col_header == 'Main Keyword':
                    cell.fill = main_fill
                    cell.font = bold_font
                elif 'Secondary Keywords' in str(col_header):
                    cell.fill = secondary_fill
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    max_length = max(max_length, cell_length)
                except Exception:
                    pass
            
            col_name = ws.cell(row=1, column=ord(column_letter) - ord('A') + 1).value
            
            if col_name == 'URL':
                ws.column_dimensions[column_letter].width = min(max_length + 2, 60)
            elif col_name and 'Secondary Keywords' in str(col_name):
                ws.column_dimensions[column_letter].width = min(max_length + 2, 80)
            else:
                ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
        
        ws.freeze_panes = 'A2'
    
    def format_onpage_sheet(self, ws):
        header_fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        yes_fill = PatternFill(start_color=self.colors['yes'], end_color=self.colors['yes'], fill_type='solid')
        no_fill = PatternFill(start_color=self.colors['no'], end_color=self.colors['no'], fill_type='solid')
        
        header_font = Font(color='FFFFFF', bold=True)
        regular_font = Font(color='000000')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        yes_no_columns = ['Title', 'Description', 'H1', 'H2', 'Content', 'Alt Text', 'Schema']
        
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = regular_font
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                col_header = ws.cell(row=1, column=cell.column).value
                if col_header in yes_no_columns:
                    if cell.value == 'YES':
                        cell.fill = yes_fill
                    elif cell.value == 'NO':
                        cell.fill = no_fill
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    max_length = max(max_length, cell_length)
                except Exception:
                    pass
            
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        ws.freeze_panes = 'A2'
    
    def calculate_summary(self):
        if not self.onpage_results:
            return {}
        
        total = len(self.onpage_results)
        elements = ['Title', 'Description', 'H1', 'H2', 'Content', 'Alt Text', 'Schema']
        
        summary = {
            'data_source': self.data_source.upper(),
            'total_keywords': total,
            'elements': {}
        }
        
        for element in elements:
            yes_count = sum(1 for r in self.onpage_results if r[element] == 'YES')
            no_count = total - yes_count
            yes_pct = (yes_count / total * 100) if total > 0 else 0
            
            summary['elements'][element] = {
                'yes': yes_count,
                'no': no_count,
                'yes_percentage': round(yes_pct, 1)
            }
        
        if 'Traffic Opportunity' in self.onpage_results[0]:
            df_results = pd.DataFrame(self.onpage_results)
            traffic_opp = pd.to_numeric(df_results['Traffic Opportunity'], errors='coerce').sum()
            summary['total_traffic_opportunity'] = int(traffic_opp)
        
        return summary
    
    def run_complete_analysis(self):
        df = self.load_data()
        if df is None:
            return {
                'success': False,
                'error': 'Failed to load or process data file'
            }
        
        self.run_onpage_analysis(df)
        excel_filename = self.create_excel_report()
        summary = self.calculate_summary()
        
        # Convert to safe strings (remove non-ASCII)
        onpage_data = []
        for result in self.onpage_results:
            onpage_data.append({
                'keyword': str(result['Keyword']),
                'url': str(result['URL']),
                'title': result['Title'],
                'description': result['Description'],
                'h1': result['H1'],
                'h2': result['H2'],
                'content': result['Content'],
                'alt_text': result['Alt Text'],
                'schema': result['Schema'],
                'position': str(result.get('Position', 'N/A')),
                'traffic_opportunity': str(result.get('Traffic Opportunity', 'N/A'))
            })
        
        mapping_df = self.create_url_keyword_mapping(self.onpage_results)
        mapping_data = []
        for _, row in mapping_df.iterrows():
            mapping_data.append({
                'URL': str(row['URL']),
                'Main Keyword': str(row['Main Keyword']),
                'Secondary Keywords (Top 6)': str(row['Secondary Keywords (Top 6)'])
            })
        
        return {
            'success': True,
            'excel_file': excel_filename,
            'data_source': self.data_source,
            'summary': summary,
            'onpage_results': onpage_data,
            'keyword_mapping': mapping_data
        }

def main():
    # Ensure UTF-8 output
    if sys.stdout.encoding != 'utf-8':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    
    if len(sys.argv) < 2:
        result = {'success': False, 'error': 'No input file provided'}
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)
    
    input_file = sys.argv[1]
    filter_option = sys.argv[2] if len(sys.argv) > 2 else 'recommended'
    output_dir = sys.argv[3] if len(sys.argv) > 3 else 'results'
    
    try:
        analyzer = SilentKeywordAnalyzer(input_file, filter_option, output_dir)
        result = analyzer.run_complete_analysis()
        
        # Output as UTF-8 JSON
        print(json.dumps(result, ensure_ascii=False))
        
    except Exception as e:
        result = {'success': False, 'error': str(e)}
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)

if __name__ == "__main__":
    main()